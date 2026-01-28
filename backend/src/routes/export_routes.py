from flask import Blueprint, send_file, request, jsonify
from openpyxl import Workbook
from io import BytesIO
from datetime import date
from sqlalchemy import extract
from dotenv import load_dotenv

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


from src.models.registry_model import Registry
from src.settings.extensions import db

load_dotenv()

export_bp = Blueprint("export", __name__, url_prefix="/registry")

COLUMNS = [
    "NOME", "DATA DE NASC.", "DATA DE ADMISS√ÉO", "DATA DA COLETA",
    "DATA ENCE.", "TEMPO DE COLETA/ADM", "DIAGN√ìSTICO", "DESFECHO",
    "Notifica√ß√£o", "DI√ÅLISE", "MATERIAL", "MICROORGANISMO", "LOCAL",

    "AMICACINA", "AMPICILINA", "AMOXILINA-CLAVULANATO", "AMP/SULBAC",
    "AZITROMICINA", "CEFALEXINA", "CEFAZOLINA", "CEFEPIME",
    "CEFTAZIDIME", "CEFOXITINA", "CEFTRIAXONE", "CEFUROXIME",
    "CIPROFLOXACINO", "CLINDAMICINA", "DAPTOMICINA", "ERITROMICINA",
    "ERTAPENEM", "GENTAMICINA", "IMIPENEM", "LEVOFLOXACINO",
    "LINEZOLIDA", "MOXIFLOXACINA", "OXACILINA", "OFLOXACINA",
    "PENICILINA", "MEROPENEM", "MINOCICLINA", "NORFLOXACINA",
    "PIPE/TAZO", "POLIMIXINA B", "RIFAMPICINA", "STREPTOMICINA",
    "TRIME/SULFA", "TEICOPLANINA", "TETRACICLINA", "TIGECICLINA",
    "VANCOMICINA", "ANFOTERICINA B", "FLUCONAZOL", "KETOCONAZOL",
    "NAZOL", "NITROFURANTOINA", "AZTREONAM", "CEFTAZIDIMA-AVIBACTAM",
    "CEFTOLOZANO-TAZOBACTAM", "CLORANFENICOL", "CEFTAROLINA",
    "COLISTINA", "AMOXICILINA", "CEFOTAXIME", "VORICONAZOL",
    "CEFTIBUFEN", "OBSERVA√á√ÉO"
]


@export_bp.route("/export", methods=["GET"])
def export_registry_excel():
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int, default=date.today().year)

    if not month or month < 1 or month > 12:
        return jsonify({"error": "M√™s inv√°lido"}), 400

    registros = (
        Registry.query
        .filter(extract("month", Registry.data_da_coleta) == month)
        .filter(extract("year", Registry.data_da_coleta) == year)
        .order_by(Registry.nome_paciente)
        .all()
    )

    if not registros:
        return jsonify({
            "message": f"N√£o existem registros para {month:02d}/{year}"
        }), 404

    wb = Workbook()
    ws = wb.active
    ws.title = f"Relat√≥rio {month:02d}-{year}"

    ws.append(COLUMNS)
    
    # üé® Estilos
    header_fill = PatternFill(
        start_color="FFC0CB",  # Rosa claro
        end_color="FFC0CB",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="000000"  # Preto
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    # üéØ Aplicar estilo no cabe√ßalho
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


    def fmt_date(d):
        return d.strftime("%d/%m/%Y") if d else ""

    for r in registros:
        ws.append([
            r.nome_paciente,
            "",  # Data nasc (n√£o existe no modelo)
            fmt_date(r.data_admissao),
            fmt_date(r.data_da_coleta),
            fmt_date(r.data_encerramento),
            r.tempo_coletar,
            r.diagnostico,
            r.desfecho,
            r.notificacao,
            r.dialise,
            r.material_coletada,
            r.microorganismo,
            r.local,

            r.amicacina,
            r.ampicilina,
            r.amoxicilina_clavulanato,
            r.amp_sulbactam,
            r.azitromicina,
            r.cefalexina,
            r.cefazolina,
            r.cefepime,
            r.ceftazidime,
            r.cefoxitina,
            r.ceftriaxone,
            r.cefuroxime,
            r.ciprofloxacino,
            r.clindamicina,
            r.daptomicina,
            r.eritromicina,
            r.ertapenem,
            r.gentamicina,
            r.imipenem,
            r.levofloxacino,
            r.linezolida,
            r.moxifloxacina,
            r.oxacilina,
            r.ofloxacina,
            r.penicilina,
            r.meropenem,
            r.minociclina,
            r.norfloxacina,
            r.piperacilina_tazobactam,
            r.polimixina_b,
            r.rifampicina,
            r.streptomicina,
            r.trimetoprim_sulfametoxazol,
            r.teicoplanina,
            r.tetraciclina,
            r.tigeciclina,
            r.vancomicina,
            r.anfotericina_b,
            r.fluconazol,
            r.ketoconazol,
            r.nazol,
            r.nitrofurantoina,
            r.aztreonam,
            r.ceftazidima_avibactam,
            r.ceftolozano_tazobactam,
            r.cloranfenicol,
            r.ceftarolina,
            r.colistina,
            r.amoxicilina,
            r.cefotaxime,
            r.voriconazol,
            r.ceftibufen,
            r.observacao
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorio_{month:02d}-{year}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
